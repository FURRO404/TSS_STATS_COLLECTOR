import aiohttp
import asyncio
import os
import json
import aiofiles
from openpyxl import Workbook, load_workbook
import time
import logging
import re
from tqdm.asyncio import tqdm

logging.basicConfig(filename='LOGS.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

CONCURRENT_REQUESTS = 5
TEAM_FETCH_CONCURRENCY = 10
TIMEOUT_DURATION = 120
RETRY_LIMIT = 5

async def get_tournament_stats(session, tournament_id, semaphore, retries=0):
    url = "https://tss.warthunder.com/functions.php"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"action": "GetStatsTournamentShort", "tournamentID": tournament_id}

    async with semaphore:
        logger.info(f"Starting data fetch for tournament ID {tournament_id}")
        start_time = time.monotonic()
        try:
            async with session.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION) as response:
                if response.status == 200:
                    try:
                        end_time = time.monotonic()
                        result = await response.json()
                        if result.get("status") == "ERROR":
                            logger.warning(f"Tournament ID {tournament_id} returned an error: {result.get('data')}")
                            return tournament_id, None  # Skip this tournament if it has an error
                        logger.info(f"Completed data fetch for tournament ID {tournament_id} in {end_time - start_time:.2f} seconds")
                        return tournament_id, result
                    except aiohttp.ContentTypeError:
                        text = await response.text()
                        try:
                            end_time = time.monotonic()
                            logger.info(f"Completed data fetch (text parse) for tournament ID {tournament_id} in {end_time - start_time:.2f} seconds")
                            return tournament_id, json.loads(text)
                        except json.JSONDecodeError:
                            logger.error(f"Failed to parse JSON for tournament ID {tournament_id}")
                else:
                    logger.error(f"Request for tournament ID {tournament_id} failed with status code {response.status}")
        except (aiohttp.ClientConnectorError, asyncio.TimeoutError) as e:
            if retries < RETRY_LIMIT:
                wait_time = 2 ** retries
                logger.warning(f"Error for tournament ID {tournament_id}: {e}. Retrying in {wait_time} seconds...")
                await asyncio.sleep(wait_time)
                return await get_tournament_stats(session, tournament_id, semaphore, retries + 1)
            logger.error(f"Failed to connect for tournament ID {tournament_id} after {RETRY_LIMIT} retries.")
    return tournament_id, None

async def save_tournament_stats(tournament_id, stats):
    start_time = time.monotonic()
    dir_path = f"TSS/{tournament_id}"
    os.makedirs(dir_path, exist_ok=True)

    async with aiofiles.open(os.path.join(dir_path, "TEAMS.json"), "w") as f:
        await f.write(json.dumps(stats, indent=4))
    logger.info(f"Saved stats for tournament ID {tournament_id}")

    output_path = await generate_output_file(tournament_id, stats)
    await fetch_team_details(tournament_id, output_path)
    await update_master_data(output_path)

    end_time = time.monotonic()
    logger.info(f"Finished processing for tournament ID {tournament_id} in {end_time - start_time:.2f} seconds")

async def generate_output_file(tournament_id, stats):
    output_data = [
        {
            "TournamentID": entry["tournamentID"],
            "TeamID": entry["teamID"],
            "TeamName": entry["realName"],
            "Place": entry["place"],
            "members": []
        }
        for entry in stats.get("readyTopTeamsTournament", [])
    ]

    output_path = f"TSS/{tournament_id}/OUTPUT.json"
    async with aiofiles.open(output_path, "w") as f:
        await f.write(json.dumps(output_data, indent=4))
    logger.info(f"Saved output for tournament ID {tournament_id}")
    return output_path

async def fetch_team_details(tournament_id, output_path):
    async with aiofiles.open(output_path, "r") as f:
        output_data = json.loads(await f.read())

    team_semaphore = asyncio.Semaphore(TEAM_FETCH_CONCURRENCY)
    async with aiohttp.ClientSession() as session:
        team_info_list = await asyncio.gather(
            *[get_team_info(session, tournament_id, team["TeamID"], team_semaphore) for team in output_data]
        )

    for team_info in team_info_list:
        if team_info:
            team_id, members, prize_pool = team_info  # Adjusted to include prize_pool
            for team in output_data:
                if team["TeamID"] == team_id:
                    team["members"] = members
                    # Add Prize field if this team has Place == 1
                    if team.get("Place") == "1" and prize_pool is not None:
                        team["Prize"] = prize_pool

    async with aiofiles.open(output_path, "w") as f:
        await f.write(json.dumps(output_data, indent=4))

async def get_team_info(session, tournament_id, team_id, team_semaphore, retries=0):
    url = "https://tss.warthunder.com/functions.php"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"action": "infoTeam", "tournamentID": tournament_id, "teamID": team_id}

    async with team_semaphore:
        try:
            async with session.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION) as response:
                if response.status == 200:
                    try:
                        team_data = await response.json()
                    except aiohttp.ContentTypeError:
                        try:
                            team_data = json.loads(await response.text())
                        except json.JSONDecodeError:
                            logger.error(f"Failed to parse JSON for team ID {team_id} in tournament ID {tournament_id}")
                            return team_id, None, None  # Return None if parsing fails

                    # Get the prize_pool string
                    prize_pool_str = team_data.get("param_tournaments", {}).get("prize_pool", "0")
                    
                    # Use regex to extract only the number before "Golden eagles", with case insensitivity and flexible spacing
                    match = re.search(r"(\d+)\s*Golden\s*eagles", prize_pool_str, re.IGNORECASE)
                    
                    # Convert to integer if a match is found, otherwise return 0
                    prize_pool = int(match.group(1)) if match else 0
                    logger.info(f"prize pool for tournament {tournament_id}: {prize_pool}")

                    # Save team data in MEMBERS folder
                    members_dir = f"TSS/{tournament_id}/MEMBERS"
                    os.makedirs(members_dir, exist_ok=True)
                    team_file_path = f"{members_dir}/{team_id}.json"
                    async with aiofiles.open(team_file_path, "w") as f:
                        await f.write(json.dumps(team_data, indent=4))

                    # Prepare member details
                    members = [{"nick": user["nick"], "userID": user["userID"]} for user in team_data.get("users_team", [])]
                    return team_id, members, prize_pool

        except (aiohttp.ClientConnectorError, asyncio.TimeoutError) as e:
            if retries < RETRY_LIMIT:
                wait_time = 2 ** retries
                logger.warning(f"Error for team ID {team_id}: {e}. Retrying in {wait_time} seconds...")
                await asyncio.sleep(wait_time)
                return await get_team_info(session, tournament_id, team_id, team_semaphore, retries + 1)
            logger.error(f"Failed to connect for team ID {team_id} after {RETRY_LIMIT} retries.")
    
    return team_id, None, None


async def update_master_data(output_path):
    master_data_path = "MASTER.json"
    
    # Load or initialize MASTER.json
    if os.path.exists(master_data_path):
        async with aiofiles.open(master_data_path, "r") as f:
            master_data = json.loads(await f.read())
    else:
        master_data = {
            "PlayerFirstPlaces": {},
            "PlayerTotalPlayed": {},
            "TeamFirstPlaces": {},
            "PlayersInTeams": {}
        }

    # Convert PlayersInTeams lists back to sets for processing
    for team, players in master_data["PlayersInTeams"].items():
        master_data["PlayersInTeams"][team] = set(players)

    # Load tournament output data
    async with aiofiles.open(output_path, "r") as f:
        output_data = json.loads(await f.read())

    # Skip updating if output_data is empty
    if not output_data:
        return

    # Update master data with the current tournament's output
    for team in output_data:
        team_name = team["TeamName"]
        place = team["Place"]
        prize = team.get("Prize", 0)  # Get prize amount if available, default to 0
        members = team.get("members", [])

        if str(place) == "1":
            # Update TeamFirstPlaces with count and prize total
            if team_name not in master_data["TeamFirstPlaces"]:
                master_data["TeamFirstPlaces"][team_name] = {"count": 0, "total_prize": 0}
            
            master_data["TeamFirstPlaces"][team_name]["count"] += 1
            master_data["TeamFirstPlaces"][team_name]["total_prize"] += prize

            for member in members:
                player_key = f"{member['nick']} ({member['userID']})"
                if player_key not in master_data["PlayerFirstPlaces"]:
                    master_data["PlayerFirstPlaces"][player_key] = {"count": 0, "total_prize": 0}

                # Increment first place count and add prize money for player
                master_data["PlayerFirstPlaces"][player_key]["count"] += 1
                master_data["PlayerFirstPlaces"][player_key]["total_prize"] += prize

        # Update total games played by each member
        for member in members:
            player_key = f"{member['nick']} ({member['userID']})"
            master_data["PlayerTotalPlayed"][player_key] = master_data["PlayerTotalPlayed"].get(player_key, 0) + 1

            if team_name not in master_data["PlayersInTeams"]:
                master_data["PlayersInTeams"][team_name] = set()
            master_data["PlayersInTeams"][team_name].add(member["nick"])

    # Convert sets back to lists for JSON compatibility
    master_data["PlayersInTeams"] = {team: list(players) for team, players in master_data["PlayersInTeams"].items()}

    async with aiofiles.open(master_data_path, "w") as f:
        await f.write(json.dumps(master_data, indent=4))

async def write_master_to_excel():
    master_data_path = "MASTER.json"
    excel_path = "MASTER.xlsx"

    async with aiofiles.open(master_data_path, "r") as f:
        master_data = json.loads(await f.read())

    workbook = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()

    def write_sheet(sheet_name, headers, data_rows):
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in data_rows:
            sheet.append(row)

    # Prepare data for PlayerFirstPlaces
    player_first_places_data = [
        [
            player.split(" (")[1][:-1],  # UserID
            player.split(" (")[0],       # Player name
            details["count"],            # FirstPlaceCount
            details["total_prize"]       # Total prize money won
        ]
        for player, details in master_data["PlayerFirstPlaces"].items()
    ]
    write_sheet("PlayerFirstPlaces", ["UserID", "Player", "FirstPlaceCount", "TotalPrize"], player_first_places_data)

    # Prepare data for PlayerTotalPlayed
    player_total_played_data = [
        [player.split(" (")[1][:-1], player.split(" (")[0], count]
        for player, count in master_data["PlayerTotalPlayed"].items()
    ]
    write_sheet("PlayerTotalPlayed", ["UserID", "Player", "TotalPlayedCount"], player_total_played_data)

    # Prepare data for TeamFirstPlaces
    team_first_places_data = [
        [team, details["count"], details["total_prize"]]
        for team, details in master_data["TeamFirstPlaces"].items()
    ]
    write_sheet("TeamFirstPlaces", ["Team", "FirstPlaceCount", "TotalPrize"], team_first_places_data)

    # Prepare data for PlayersInTeams with the desired format
    players_in_teams_data = []
    for team, players in master_data["PlayersInTeams"].items():
        # Add the team name only once, then list each player in a new row under it
        players_in_teams_data.append([team, ""])
        players_in_teams_data.extend([["", player] for player in players])

    write_sheet("PlayersInTeams", ["Team", "Players"], players_in_teams_data)

    # Remove the default sheet if it exists
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    workbook.save(excel_path)
    logger.info(f"Data successfully written to {excel_path}")

min = 0
max = 20900             #EDIT THIS TO WHATEVER YOU NEED

async def main():
    async with aiohttp.ClientSession() as session:
        semaphore = asyncio.Semaphore(CONCURRENT_REQUESTS)
        
        tournament_ids = range(min, max)
        tasks = [
            get_tournament_stats(session, tournament_id, semaphore)
            for tournament_id in tournament_ids
            if not os.path.exists(f"TSS/{tournament_id}/OUTPUT.json")
        ]

        # Use tqdm to show a progress bar in the console
        for task in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Processing Tournaments", unit="tournament"):
            tid, stats = await task
            if stats:
                await save_tournament_stats(tid, stats)
                
        await write_master_to_excel()

asyncio.run(main())
